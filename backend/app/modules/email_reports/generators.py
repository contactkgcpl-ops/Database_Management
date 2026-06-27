import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Styles definition
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="000000")
BOLD_FONT = Font(name="Calibri", size=11, bold=True, color="000000")
NORMAL_FONT = Font(name="Calibri", size=11, color="000000")
SMALL_FONT = Font(name="Calibri", size=9, italic=True, color="555555")

BLUE_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
GREEN_FILL = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
PURPLE_FILL = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")

LIGHT_BLUE_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LIGHT_ORANGE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
LIGHT_PURPLE_FILL = PatternFill(start_color="F2EBF7", end_color="F2EBF7", fill_type="solid")

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin", color="D3D3D3"),
    right=Side(style="thin", color="D3D3D3"),
    top=Side(style="thin", color="D3D3D3"),
    bottom=Side(style="thin", color="D3D3D3")
)

DOUBLE_BOTTOM_BORDER = Border(
    top=Side(style="thin", color="000000"),
    bottom=Side(style="double", color="000000")
)

def auto_fit_columns(ws, min_width=12):
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if '\n' in val_str:
                max_len = max(max_len, max(len(p) for p in val_str.split('\n')))
            else:
                max_len = max(max_len, len(val_str))
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, min_width)

def generate_excel_report(data: dict, target_date: date) -> bytes:
    wb = Workbook()
    
    # ------------------ SHEET 1: DAILY SUMMARY ------------------
    ws1 = wb.active
    ws1.title = "Daily Summary"
    ws1.views.sheetView[0].showGridLines = True
    
    current_row = 1
    
    # Page Title
    ws1.cell(row=current_row, column=1, value=f"Daily Activity Summary Report — {target_date.strftime('%d-%b-%Y')}").font = TITLE_FONT
    current_row += 2
    
    # 🔵 SECTION A: Company + User Summary
    ws1.cell(row=current_row, column=1, value="🔵 Section A: Company + User Daily Activity Summary").font = BOLD_FONT
    current_row += 1
    
    headers_a = ["Our Company / User", "Assigned Data", "Worked", "Pending Data", "Calls", "WhatsApp", "Email", "Social Media"]
    for col_idx, h in enumerate(headers_a, 1):
        cell = ws1.cell(row=current_row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = BLUE_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        
    current_row += 1
    start_row_a = current_row
    
    for oc in data["oc_summary"]:
        # Parent OurCompany row
        row_cells = [
            ws1.cell(row=current_row, column=1, value=oc["name"]),
            ws1.cell(row=current_row, column=2, value=oc["assigned"]),
            ws1.cell(row=current_row, column=3, value=oc["worked"]),
            ws1.cell(row=current_row, column=4, value=oc["pending"]),
            ws1.cell(row=current_row, column=5, value=oc["calls"]),
            ws1.cell(row=current_row, column=6, value=oc["whatsapp"]),
            ws1.cell(row=current_row, column=7, value=oc["email"]),
            ws1.cell(row=current_row, column=8, value=oc["social"]),
        ]
        for idx, cell in enumerate(row_cells):
            cell.font = BOLD_FONT
            cell.fill = LIGHT_BLUE_FILL
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN if idx == 0 else RIGHT_ALIGN
            
        current_row += 1
        
        # Children User rows
        for u_row in oc["users"]:
            row_cells_u = [
                ws1.cell(row=current_row, column=1, value=u_row["name"]),
                ws1.cell(row=current_row, column=2, value=u_row["assigned"]),
                ws1.cell(row=current_row, column=3, value=u_row["worked"]),
                ws1.cell(row=current_row, column=4, value=u_row["pending"]),
                ws1.cell(row=current_row, column=5, value=u_row["calls"]),
                ws1.cell(row=current_row, column=6, value=u_row["whatsapp"]),
                ws1.cell(row=current_row, column=7, value=u_row["email"]),
                ws1.cell(row=current_row, column=8, value=u_row["social"]),
            ]
            for idx, cell in enumerate(row_cells_u):
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
                cell.alignment = LEFT_ALIGN if idx == 0 else RIGHT_ALIGN
            current_row += 1
            
    # Section A Grand Total
    end_row_a = current_row - 1
    sum_assigned = sum(oc["assigned"] for oc in data["oc_summary"])
    sum_worked = sum(oc["worked"] for oc in data["oc_summary"])
    sum_pending = sum(oc["pending"] for oc in data["oc_summary"])
    sum_calls = sum(oc["calls"] for oc in data["oc_summary"])
    sum_wa = sum(oc["whatsapp"] for oc in data["oc_summary"])
    sum_email = sum(oc["email"] for oc in data["oc_summary"])
    sum_social = sum(oc["social"] for oc in data["oc_summary"])
    
    total_vals = ["GRAND TOTAL", sum_assigned, sum_worked, sum_pending, sum_calls, sum_wa, sum_email, sum_social]
    for idx, val in enumerate(total_vals):
        cell = ws1.cell(row=current_row, column=idx+1, value=val)
        cell.font = BOLD_FONT
        cell.border = DOUBLE_BOTTOM_BORDER
        cell.alignment = LEFT_ALIGN if idx == 0 else RIGHT_ALIGN
        
    current_row += 3
    
    # 🟢 SECTION B: User Time Log
    ws1.cell(row=current_row, column=1, value="🟢 Section B: User Daily Time Log & Availability Status").font = BOLD_FONT
    current_row += 1
    
    headers_b = ["User Name", "Login Time", "Logout Time", "Break Time", "Working Hours", "Availability Status"]
    for col_idx, h in enumerate(headers_b, 1):
        cell = ws1.cell(row=current_row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = GREEN_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        
    current_row += 1
    for log in data["user_time_logs"]:
        row_cells = [
            ws1.cell(row=current_row, column=1, value=log["name"]),
            ws1.cell(row=current_row, column=2, value=log["login"]),
            ws1.cell(row=current_row, column=3, value=log["logout"]),
            ws1.cell(row=current_row, column=4, value=log["break"]),
            ws1.cell(row=current_row, column=5, value=log["work"]),
            ws1.cell(row=current_row, column=6, value=log["status"]),
        ]
        for idx, cell in enumerate(row_cells):
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if idx == 5: # Availability Status
                cell.font = BOLD_FONT
                # Color status values
                status_str = str(log["status"])
                if "Active" in status_str:
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # light green
                elif "On Leave" in status_str:
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # light yellow
                elif "Unavailable" in status_str:
                    cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # light orange
                else:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # light gray
            cell.alignment = CENTER_ALIGN if idx > 0 else LEFT_ALIGN
        current_row += 1
        
    current_row += 3
    
    # 🟠 SECTION C: Pending Work Summary
    ws1.cell(row=current_row, column=1, value="🟠 Section C: User Daily Pending Work Summary").font = BOLD_FONT
    current_row += 1
    
    headers_c = ["User Name", "Our Company", "Assigned Today", "Worked Today", "Pending", "Pending %"]
    for col_idx, h in enumerate(headers_c, 1):
        cell = ws1.cell(row=current_row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = ORANGE_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        
    for p in data["pending_summary"]:
        row_cells = [
            ws1.cell(row=current_row, column=1, value=p["user_name"]),
            ws1.cell(row=current_row, column=2, value=p["our_company"]),
            ws1.cell(row=current_row, column=3, value=p["assigned"]),
            ws1.cell(row=current_row, column=4, value=p["worked"]),
            ws1.cell(row=current_row, column=5, value=p["pending"]),
            ws1.cell(row=current_row, column=6, value=p["pending_pct"]),
        ]
        is_bold = p.get("is_bold", False)
        for idx, cell in enumerate(row_cells):
            cell.font = BOLD_FONT if is_bold else NORMAL_FONT
            cell.border = THIN_BORDER
            if is_bold:
                cell.fill = LIGHT_ORANGE_FILL
            cell.alignment = CENTER_ALIGN if idx >= 2 else LEFT_ALIGN
        current_row += 1
        
    current_row += 3
    
    # 🟣 SECTION D: City + Industry Breakdown (Combined)
    ws1.cell(row=current_row, column=1, value="🟣 Section D: City + Industry Activity Breakdown").font = BOLD_FONT
    current_row += 1
    
    headers_d = ["City / Industry", "Calls", "WhatsApp", "Email", "Social Media", "Total Activity"]
    for col_idx, h in enumerate(headers_d, 1):
        cell = ws1.cell(row=current_row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = PURPLE_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        
    current_row += 1
    for ci in data["city_industry"]:
        row_cells = [
            ws1.cell(row=current_row, column=1, value=ci["key"]),
            ws1.cell(row=current_row, column=2, value=ci["calls"]),
            ws1.cell(row=current_row, column=3, value=ci["whatsapp"]),
            ws1.cell(row=current_row, column=4, value=ci["email"]),
            ws1.cell(row=current_row, column=5, value=ci["social"]),
            ws1.cell(row=current_row, column=6, value=ci["total"]),
        ]
        for idx, cell in enumerate(row_cells):
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = RIGHT_ALIGN if idx > 0 else LEFT_ALIGN
        current_row += 1
        
    auto_fit_columns(ws1, min_width=15)

    # ─── SECTION E: User Company + City + Industry Tracking ────────────────────
    current_row += 3
    ws1.cell(row=current_row, column=1, value="🔴 Section E: User-wise Our Company + City + Industry Data Tracking").font = BOLD_FONT
    current_row += 1

    RED_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    LIGHT_RED_FILL = PatternFill(start_color="FFE7E7", end_color="FFE7E7", fill_type="solid")

    headers_e = ["User Name", "Our Company", "Assignee", "City", "Industry", "Assigned", "Worked", "Pending", "Calls", "WhatsApp", "Email", "Social"]
    for col_idx, h in enumerate(headers_e, 1):
        cell = ws1.cell(row=current_row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = RED_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    current_row += 1
    prev_user = None
    for row in data.get("user_company_tracking", []):
        is_new_user = row["user_name"] != prev_user
        row_cells = [
            ws1.cell(row=current_row, column=1, value=row["user_name"] if is_new_user else ""),
            ws1.cell(row=current_row, column=2, value=row["our_company"]),
            ws1.cell(row=current_row, column=3, value=row["assignee"]),
            ws1.cell(row=current_row, column=4, value=row["city"]),
            ws1.cell(row=current_row, column=5, value=row["industry"]),
            ws1.cell(row=current_row, column=6, value=row["assigned"]),
            ws1.cell(row=current_row, column=7, value=row["worked"]),
            ws1.cell(row=current_row, column=8, value=row["pending"]),
            ws1.cell(row=current_row, column=9, value=row["calls"]),
            ws1.cell(row=current_row, column=10, value=row["whatsapp"]),
            ws1.cell(row=current_row, column=11, value=row["email"]),
            ws1.cell(row=current_row, column=12, value=row["social"]),
        ]
        for idx, cell in enumerate(row_cells):
            cell.font = BOLD_FONT if is_new_user and idx == 0 else NORMAL_FONT
            cell.border = THIN_BORDER
            if is_new_user:
                cell.fill = LIGHT_RED_FILL
            cell.alignment = LEFT_ALIGN if idx <= 4 else RIGHT_ALIGN
        current_row += 1
        prev_user = row["user_name"]

    auto_fit_columns(ws1, min_width=15)
    
    # ------------------ SHEET 2+: USER WISE REPORTS ------------------
    for ur in data["per_user_reports"]:
        sheet_name = f"User - {ur['user_name']}"[:31]
        ws = wb.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True
        
        ws.cell(row=1, column=1, value=f"Daily Performance Report: {ur['user_name']}").font = TITLE_FONT
        ws.cell(row=2, column=1, value=f"Report Date: {target_date.strftime('%d-%b-%Y')}").font = SMALL_FONT
        
        ws.cell(row=4, column=1, value="Daily Activity Totals").font = BOLD_FONT
        sum_headers = ["Assigned Data", "Worked Today", "Pending Data", "Calls", "WhatsApp", "Email", "Social Media"]
        for col_idx, h in enumerate(sum_headers, 1):
            cell = ws.cell(row=5, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = BLUE_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            
        t = ur["totals"]
        sum_vals = [t["assigned"], t["worked"], t["pending"], t["calls"], t["whatsapp"], t["email"], t["social"]]
        for col_idx, val in enumerate(sum_vals, 1):
            cell = ws.cell(row=6, column=col_idx, value=val)
            cell.font = BOLD_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            
        ws.cell(row=8, column=1, value="City / Industry Breakdown").font = BOLD_FONT
        ci_headers = ["City / Industry", "Calls", "WhatsApp", "Email", "Social Media", "Total Activity"]
        for col_idx, h in enumerate(ci_headers, 1):
            cell = ws.cell(row=9, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = PURPLE_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            
        cur_r = 10
        for ci in ur["city_industry"]:
            row_cells = [
                ws.cell(row=cur_r, column=1, value=ci["key"]),
                ws.cell(row=cur_r, column=2, value=ci["calls"]),
                ws.cell(row=cur_r, column=3, value=ci["whatsapp"]),
                ws.cell(row=cur_r, column=4, value=ci["email"]),
                ws.cell(row=cur_r, column=5, value=ci["social"]),
                ws.cell(row=cur_r, column=6, value=ci["total"]),
            ]
            for idx, cell in enumerate(row_cells):
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
                cell.alignment = RIGHT_ALIGN if idx > 0 else LEFT_ALIGN
            cur_r += 1
            
        cur_r += 2
        ws.cell(row=cur_r, column=1, value="Worked Companies Detail List").font = BOLD_FONT
        cur_r += 1
        
        detail_headers = ["#", "Company Name", "Contact Number", "City", "Industry", "Connected Via", "Response / Status", "Last Activity Time", "Assigned By"]
        for col_idx, h in enumerate(detail_headers, 1):
            cell = ws.cell(row=cur_r, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = GREEN_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            
        cur_r += 1
        for idx, wl in enumerate(ur["worked_details"], 1):
            row_cells = [
                ws.cell(row=cur_r, column=1, value=idx),
                ws.cell(row=cur_r, column=2, value=wl["company_name"]),
                ws.cell(row=cur_r, column=3, value=wl["contact_number"]),
                ws.cell(row=cur_r, column=4, value=wl["city"]),
                ws.cell(row=cur_r, column=5, value=wl["industry"]),
                ws.cell(row=cur_r, column=6, value=wl["connected_via"]),
                ws.cell(row=cur_r, column=7, value=wl["response"]),
                ws.cell(row=cur_r, column=8, value=wl["last_activity_time"]),
                ws.cell(row=cur_r, column=9, value=wl["assigned_by"]),
            ]
            for c_idx, cell in enumerate(row_cells):
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
                if c_idx == 0:
                    cell.alignment = CENTER_ALIGN
                elif c_idx in (1, 2):
                    cell.alignment = LEFT_ALIGN
                else:
                    cell.alignment = CENTER_ALIGN if c_idx in (5, 7) else LEFT_ALIGN
            cur_r += 1
            
        cur_r += 2
        ws.cell(row=cur_r, column=1, value="Pending Companies Detail List").font = BOLD_FONT
        cur_r += 1
        
        pending_headers = ["#", "Company Name", "Contact Number", "City", "Industry", "Assigned Time", "Days Pending"]
        for col_idx, h in enumerate(pending_headers, 1):
            cell = ws.cell(row=cur_r, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = ORANGE_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            
        cur_r += 1
        for idx, pl in enumerate(ur["pending_details"], 1):
            row_cells = [
                ws.cell(row=cur_r, column=1, value=idx),
                ws.cell(row=cur_r, column=2, value=pl["company_name"]),
                ws.cell(row=cur_r, column=3, value=pl["contact_number"]),
                ws.cell(row=cur_r, column=4, value=pl["city"]),
                ws.cell(row=cur_r, column=5, value=pl["industry"]),
                ws.cell(row=cur_r, column=6, value=pl["assigned_at"]),
                ws.cell(row=cur_r, column=7, value=pl["days_pending"]),
            ]
            for c_idx, cell in enumerate(row_cells):
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
                if c_idx in (0, 6):
                    cell.alignment = CENTER_ALIGN
                else:
                    cell.alignment = LEFT_ALIGN
            cur_r += 1
            
        auto_fit_columns(ws, min_width=12)
        
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()
